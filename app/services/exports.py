from typing import Any, Dict, List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter




# ---------------------------
# Helpers de formato Excel
# ---------------------------
def _autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            try:
                val = "" if cell.value is None else str(cell.value)
            except Exception:
                val = ""
            max_length = max(max_length, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)

def _header(ws, row: int, values: List[str]) -> None:
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")


# ---------------------------
# Armado del Excel
# ---------------------------
def build_excel_from_liquidacion(payload: Dict[str, Any]) -> bytes:
    """
    Espera un JSON con las claves:
      status, solicitud{obra_sociales[], periodos_normalizados[]},
      resumen{...},
      por_medico: [ {medico_id, medico_nombre, obras_sociales: [{obra_social, periodos: [{periodo, totales{...}, prestaciones: [...] }]}]} ]
    """
    # Validación mínima
    if payload.get("status") != "ok":
        raise ValueError(payload.get("message") or "Payload inválido (status != ok)")

    por_medico: List[Dict[str, Any]] = payload.get("por_medico", [])
    solicitud: Dict[str, Any] = payload.get("solicitud", {})
    resumen: Dict[str, Any] = payload.get("resumen", {})

    wb = Workbook()

    # 1) Hoja Resumen
    ws1 = wb.active
    ws1.title = "Resumen"
    _header(ws1, 1, ["Campo", "Valor"])

    fila = 2
    ws1.cell(fila, 1, "Obras sociales")
    ws1.cell(fila, 2, ", ".join(solicitud.get("obra_sociales", []))); fila += 1

    ws1.cell(fila, 1, "Períodos")
    ws1.cell(fila, 2, ", ".join(solicitud.get("periodos_normalizados", []))); fila += 1

    # Totales
    for k_print, k_key in [
        ("Total prestaciones incluidas", "total_prestaciones_incluidas"),
        ("Total bruto", "total_bruto"),
        ("Total descuentos", "total_descuentos"),
        ("Total neto", "total_neto"),
    ]:
        ws1.cell(fila, 1, k_print)
        ws1.cell(fila, 2, resumen.get(k_key, 0)); fila += 1

    _autosize(ws1)

    # 2) Hoja Detalle por médico  (una fila por médico→OS→periodo con totales del periodo)
    ws2 = wb.create_sheet("Detalle por médico")
    _header(ws2, 1, [
        "Médico ID", "Médico", "Obra social", "Período",
        "Bruto periodo", "Descuentos periodo", "Neto periodo"
    ])

    fila = 2
    for medico in por_medico:
        medico_id = medico.get("medico_id")
        medico_nombre = medico.get("medico_nombre")
        for os_block in medico.get("obras_sociales", []):
            os_name = os_block.get("obra_social")
            for periodo_block in os_block.get("periodos", []):
                periodo = periodo_block.get("periodo")
                tot = periodo_block.get("totales", {}) or {}
                bruto = round(float(tot.get("bruto", 0) or 0), 2)
                descuentos = round(float(tot.get("descuentos", 0) or 0), 2)
                neto = round(float(tot.get("neto", 0) or 0), 2)

                ws2.cell(fila, 1, medico_id)
                ws2.cell(fila, 2, medico_nombre)
                ws2.cell(fila, 3, os_name)
                ws2.cell(fila, 4, periodo)
                ws2.cell(fila, 5, bruto)
                ws2.cell(fila, 6, descuentos)
                ws2.cell(fila, 7, neto)
                fila += 1

    _autosize(ws2)

    # 3) Hoja Prestaciones (cada prestación en una fila)
    ws3 = wb.create_sheet("Prestaciones")
    _header(ws3, 1, [
        "Médico ID", "Médico",
        "Obra social", "Período",
        "Atención ID", "Código prestación", "Fecha",
        "Bruto", "Descuentos", "Neto"
    ])
    fila = 2

    for medico in por_medico:
        medico_id = medico.get("medico_id")
        medico_nombre = medico.get("medico_nombre")
        for os_block in medico.get("obras_sociales", []):
            os_name = os_block.get("obra_social")
            for periodo_block in os_block.get("periodos", []):
                periodo = periodo_block.get("periodo")
                for p in periodo_block.get("prestaciones", []):
                    id_atencion = p.get("id_atencion")
                    codigo = p.get("codigo_prestacion")
                    fecha = p.get("fecha")
                    bruto = round(float(p.get("bruto", 0) or 0), 2)
                    descuentos = round(float(p.get("descuentos", 0) or 0), 2)
                    neto = round(float(p.get("neto", bruto - descuentos) or (bruto - descuentos)), 2)

                    ws3.cell(fila, 1, medico_id)
                    ws3.cell(fila, 2, medico_nombre)
                    ws3.cell(fila, 3, os_name)
                    ws3.cell(fila, 4, periodo)
                    ws3.cell(fila, 5, id_atencion)
                    ws3.cell(fila, 6, codigo)
                    ws3.cell(fila, 7, fecha)
                    ws3.cell(fila, 8, bruto)
                    ws3.cell(fila, 9, descuentos)
                    ws3.cell(fila, 10, neto)
                    fila += 1

    _autosize(ws3)

    # Guardar a memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()